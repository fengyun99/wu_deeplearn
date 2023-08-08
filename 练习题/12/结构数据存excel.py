#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2023/8/1 14:47
# @Author   : FengYun
# @File     : 结构数据存excel.py
# @Software : PyCharm
'''
将提供的结构化数据存储到excel当中
'''
from openpyxl import load_workbook, Workbook


def data_to_excel(excel_name, data):
    # 创建xlsx文件
    wb = Workbook()
    wb.save(excel_name)

    # 加载Excel文件
    wb = load_workbook(excel_name)

    # 选择默认的工作表（第一个工作表）
    ws = wb.active
    temp = []
    if isinstance(data,dict):
        for k, v in data.items():
            res = []
            if isinstance(v, list):
                res.append(k)
                res.extend(v)
                temp.append(res)
            else:
                res = [k, v]
                temp.append(res)
    elif isinstance(data,list):
        for i in data:
            temp.append(i)

    length_data = len(temp[0])
    length_temp = len(temp)
    for i in range(length_temp):
        for j in range(length_data):
            ws.cell(row=i + 1, column=j + 1).value = temp[i][j]

    wb.save(excel_name)


if __name__ == '__main__':
    data01 = {
        "1": ["张三", 150, 120, 100],
        "2": ["李四", 90, 99, 95],
        "3": ["王五", 60, 66, 68]
    }
    data_to_excel("example.xlsx", data01)
    data02 = {
        "1": "上海",
        "2": "北京",
        "3": "成都"
    }
    data_to_excel("city.xlsx", data02)
    data03 = [
        [1, 82, 65535],
        [20, 90, 13],
        [26, 809, 1024]
    ]
    data_to_excel("numbers.xlsx", data03)
